
# list_all_ids_to_xlsx.py
# ------------------------------------------------------------
# Scans PACK JSON files and exports an Excel workbook:
# - One sheet per module (DNAIND, DNAI, AGGLUT, BLOODG, ...)
# - Also an "ALL" sheet.
#
# It scans:
#  1) ./output/*_PACK_*.json (recommended)
#  2) ./packs/*.json (your DNAIND no-figure style)
#     - If a file is an "item pack" (dict with module/items), it is read.
#     - If it's a list of items (older style), it is read.
#     - If it's single problem dict with "id"/"module", it is included.
#
# Run:
#   pip install openpyxl
#   python list_all_ids_to_xlsx.py
# ------------------------------------------------------------

import os, json, glob, time
from collections import defaultdict
from typing import Any, Dict, List, Tuple

OUT_XLSX = os.path.join("output", f"ALL_ID_INDEX_{int(time.time())}.xlsx")

def _safe_load(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _extract_items(obj: Any, src_path: str) -> List[Dict[str, Any]]:
    """
    Normalize to list of dict items: {id,module,qnum,seed,src}
    """
    items: List[Dict[str, Any]] = []

    if isinstance(obj, dict) and "items" in obj and isinstance(obj["items"], list):
        module = obj.get("module", "UNKNOWN")
        batch_id = obj.get("batch_id", "")
        for it in obj["items"]:
            if not isinstance(it, dict): 
                continue
            items.append({
                "id": it.get("id", ""),
                "module": it.get("module", module),
                "qnum": it.get("qnum", ""),
                "seed": it.get("seed", ""),
                "batch_id": batch_id,
                "src": os.path.basename(src_path),
            })
        return items

    if isinstance(obj, list):
        for it in obj:
            if isinstance(it, dict) and ("id" in it or "problem_id" in it):
                items.append({
                    "id": it.get("id", it.get("problem_id", "")),
                    "module": it.get("module", "UNKNOWN"),
                    "qnum": it.get("qnum", ""),
                    "seed": it.get("seed", ""),
                    "batch_id": "",
                    "src": os.path.basename(src_path),
                })
        return items

    if isinstance(obj, dict) and ("id" in obj or "problem_id" in obj):
        items.append({
            "id": obj.get("id", obj.get("problem_id", "")),
            "module": obj.get("module", "UNKNOWN"),
            "qnum": obj.get("qnum", ""),
            "seed": obj.get("seed", ""),
            "batch_id": obj.get("batch_id", ""),
            "src": os.path.basename(src_path),
        })
        return items

    # unknown
    return items

def collect_all_items() -> List[Dict[str, Any]]:
    paths = []
    paths += glob.glob(os.path.join("output", "*_PACK_*.json"))
    paths += glob.glob(os.path.join("packs", "*.json"))     # old per-problem or ALL_PACKS
    paths += glob.glob(os.path.join("packs", "*.jsonl"))    # ignore (not supported)
    items = []
    for p in sorted(set(paths)):
        if p.endswith(".jsonl"):
            continue
        try:
            obj = _safe_load(p)
        except Exception:
            continue
        items.extend(_extract_items(obj, p))
    # filter empties
    items = [it for it in items if it.get("id")]
    return items

def write_xlsx(items: List[Dict[str, Any]]):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    os.makedirs("output", exist_ok=True)
    wb = Workbook()
    # default sheet -> ALL
    ws_all = wb.active
    ws_all.title = "ALL"

    headers = ["module", "id", "qnum", "seed", "batch_id", "src"]
    ws_all.append(headers)
    for it in items:
        ws_all.append([it.get(h, "") for h in headers])

    # sheets by module
    by_mod = defaultdict(list)
    for it in items:
        by_mod[it.get("module", "UNKNOWN")].append(it)

    for mod, rows in sorted(by_mod.items(), key=lambda x: x[0]):
        # Excel sheet name limit 31
        name = mod[:31] if mod else "UNKNOWN"
        if name in wb.sheetnames:
            continue
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for it in rows:
            ws.append([it.get(h, "") for h in headers])

    # auto width (simple)
    for ws in wb.worksheets:
        for col_idx, h in enumerate(headers, start=1):
            max_len = max([len(str(h))] + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row+1)])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(55, max(10, max_len + 2))

    wb.save(OUT_XLSX)
    print("✅ 저장:", OUT_XLSX)

def main():
    items = collect_all_items()
    print(f"스캔된 ID 개수: {len(items)}")
    write_xlsx(items)

if __name__ == "__main__":
    main()
